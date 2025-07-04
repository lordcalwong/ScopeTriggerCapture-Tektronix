# Power Monitoring- Synchronous
#
# Continuous monitor Amplifier Output channels and/or Line Inputs with
# synchronous logging. Trigger mode should be in autorun.
#
# User inputs IP address, sample time in seconds with default of 5 seconds,
# and number of channels to monitor (1-8).
# 
# User has option to set up scope or allow continous channels to be configured
# for RMS measurements.
#
# Currently, the maximum voltage is set to 50Vp or about 300W/ch.
# Uses pyvisa for generic scope SCPI communications for both DPO4k and MSO58
# series scopes.
#
# Saves data to csv file, closes file, and imports csv into MS Excel file 
# and plots a chart.
#
# Author: C. Wong 20250703

import time
import datetime
import os
import keyboard
import pyvisa
import threading
import csv

from openpyxl import Workbook
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import Paragraph, CharacterProperties, Font
from openpyxl.styles import Font as ExcelFont

DEFAULT_IP_ADDRESS = '192.168.1.53'  #default IP, 192.168.1.53, 10.101.100.151
MIN_ACQUISITION_INTERVAL = 5   # default sampling rate
MAX_VRMS = 50

# Find user esktop one level down from home (~/* /Desktop) and set up as optional save path
from glob import glob
DESKTOP = glob(os.path.expanduser("~\\*\\Desktop"))
# If not found, revert to the standard location (~/Desktop)
if len(DESKTOP) == 0:
    DESKTOP = os.path.expanduser("~\\Desktop")
else:
    DESKTOP = DESKTOP[0] # glob returns a list, take the first result

# Global flag to signal the main loop and threads to stop
stop_program_event = threading.Event()

# --- FUNCTION DEFINITION ---
def timer_thread_func(event_to_set: threading.Event, interval: float, stop_event: threading.Event):
    """
    Separate thread that signals an event after the specified interval.
    """
    while not stop_event.is_set():
        # Wait for the interval, but allow to be interrupted by stop_event
        # If stop_event is set during this wait, it will return True immediately.
        if stop_event.wait(interval):
            break # Exit the loop if stop_event was set
        # If the wait completed (meaning interval passed and stop_event wasn't set),
        # then set the event for the main thread.
        if not stop_event.is_set(): 
            event_to_set.set()

def on_q_press():
    """
    Callback function when 'q' is pressed.
    """
    print("\n'q' pressed. Signaling program to stop.")
    stop_program_event.set()

def connect_to_instrument(resource_manager: pyvisa.ResourceManager, default_ip: str = DEFAULT_IP_ADDRESS):
    """
    Prompts the user for an IP address and attempts to establish a connection
    to a PyVISA instrument, retrying until successful.

    Args:
        resource_manager: The PyVISA ResourceManager instance.
        default_ip: The default IP address to suggest to the user.

    Returns:
        The connected PyVISA instrument object.
    """
    my_instrument = None
    while my_instrument is None:
        ip_address_input = input(
            f"Enter the instrument's IP address or 'd' for default ({default_ip}): "
        ).strip()
        if ip_address_input.lower() == 'd':
            visa_address = default_ip
        else:
            visa_address = ip_address_input

        # Construct the full VISA resource string
        resource_string = f'TCPIP::{visa_address}::INSTR'

        print(f"Attempting to connect to: {resource_string}")

        try:
            # Attempt to open the resource
            my_instrument = resource_manager.open_resource(resource_string)

            # Try to query the instrument to verify connection
            print(f"Successfully connected! Instrument ID: {my_instrument.query('*IDN?').strip()}")

        except pyvisa.errors.VisaIOError as e:
            print(f"Connection failed: {e}")
            print("Please ensure the IP address is correct and the instrument is on and connected to the network.")
            print("Retrying in 2 seconds...")
            my_instrument = None   # Ensure my_instrument is None to continue the loop
            time.sleep(2)   # Wait before retrying

        except Exception as e:
            print(f"An unexpected error occurred: {e}")
            print("Retrying in 2 seconds...")
            my_instrument = None
            time.sleep(2)

    return my_instrument

def get_num_channels():
    """
    Asks the user to input the number of channels (1-8).
    """
    while True:
        try:
            num_channels = int(input("Enter the number of channels to monitor (1-8): "))
            if 1 <= num_channels <= 8:
                return num_channels
            else:
                print("Invalid input. Please enter a number between 1 and 8.")
        except ValueError:
            print("Invalid input. Please enter a number.")

def sample_period(default_sample_time: str = MIN_ACQUISITION_INTERVAL):
    """
    Asks user to input the time between samples.
    """
    while True:
        try:
            sample_period = input(f"Enter time between samples in seconds (1-300) or 'd' for default ({default_sample_time}) : ").strip()
            if sample_period.lower() == 'd':
                return default_sample_time
            else:
                period = int(sample_period)
                if 1 <= period <= 300:
                    return period
                else:
                    print("Invalid input. Please enter a number between 1 and 300.")
        except ValueError:
            print("Invalid input. Please enter a number.")

def setup_scope(scope_device, num_channels):
    """
    Configures channels for the specified number of channels.
    Minimally tries to set scale and position.
    """
    print("Ok. Setting up oscilloscope...", end='')
    
    for i in range(1, num_channels + 1):
        scope_device.write(f"SELect:CH{i} ON")
        scope_device.write(f"CH{i}:SCALe 10")
        scope_device.write(f"CH{i}:POSition 0")
        scope_device.write(f"MEASUrement:MEAS{i}:SOUrce CH{i}; STATE 1")   # Need separate STATE 1 command for DPO 
        scope_device.write(f"MEASUrement:MEAS{i}:SOUrce CH{i}; TYPE RMS")  # Need separate TYPE for MSO 
 
    # Wait for scope to finish setting up
    scope_device.query("*OPC?")
    print("Scope setup complete.")

def make_datafile(num_channels, timestamp, desktoppath: str = DESKTOP):
    """Generates a data file for the data based on start date and time.
    Includes headers for each monitored channel. Asks user for directory
    or defaults to desktop.
    Returns tuple for user_path and data_log_file_name
    """
    while True:
        try:
            user_path_input = input(f"Enter path for data or 'd' for default ({desktoppath}): ").strip()
            if user_path_input.lower() == 'd':
                user_path = desktoppath
            else:
                user_path = user_path_input

            data_log_file_name = timestamp.strftime("%Y%m%d_%H%M%S.txt")
            full_data_path = os.path.join(user_path, data_log_file_name)

            # Check if data file already exists
            if not os.path.exists(full_data_path):
                print(f"Creating new data file: {full_data_path}")
                with open(full_data_path, "w") as datafile:
                    header = "Count, Time"
                    for i in range(1, num_channels + 1):
                        header += f", Vrms_CH{i}"
                    datafile.write(header + "\n")
            else:
                print(f"File exist. We will be appending to existing file: {full_data_path}")
            return user_path, data_log_file_name

        except ValueError:
            print("Invalid path. Please enter a valid path.")

def add_sample_to_file(save_directory, data_file_name, counter, time_in_seconds, v_rms_values):
    """
    Appends sample data for all channels to the specified file.
    """
    try:
        datafile_and_path = os.path.join(save_directory, data_file_name)
        with open(datafile_and_path, "a") as f:
            # Start with count and time
            line = f"{counter:4d}, {time_in_seconds:9.3f}"
            # Add each Vrms value
            for v_rms in v_rms_values:
                line += f", {v_rms:6.3f}"
            f.write(line + "\n")
    except IOError as e:
        print(f"Error appending data to file '{datafile_and_path}': {e}")

def apply_vrms_bounds(v_rms):
    """
    Applies upper and lower bounds to the Vrms reading.
    """
    return max(min(v_rms, MAX_VRMS), 0)

def write_to_excel_with_chart(datafile_name: str, save_directory: str, num_channels: int):
    """
    Reads data from the specified CSV file, writes it to an Excel worksheet,
    and creates a scatter chart.

    Args:
        datafile_name: The name of the CSV data file (e.g., "20250618_120000.txt").
        save_directory: The directory where the CSV and Excel files are saved.
        num_channels: The number of Vrms channels recorded.
    """
    full_csv_path = os.path.join(save_directory, datafile_name)
    excel_file_name = os.path.splitext(datafile_name)[0] + ".xlsx"
    full_excel_path = os.path.join(save_directory, excel_file_name)

    print(f"\nAttempting to create Excel file: {full_excel_path}")

    wb = Workbook()
    ws = wb.active
    ws.title = "Power Monitoring Data"

    try:
        with open(full_csv_path, 'r') as f:
            reader = csv.reader(f)
            # Write header row
            header = next(reader)
            ws.append(header)

            # Write data rows and convert to numbers, then format cells
            row_count = 0 # To track actual row in Excel, starting from 1 for headers
            for row in reader:
                row_count += 1 # Increment for data rows
                processed_row = []
                for i, value in enumerate(row):
                    if i == 0:  # 'Count' column
                        try:
                            num_value = int(value)
                            processed_row.append(num_value)
                            # Set number format for 'Count' column (optional, usually general is fine)
                            ws.cell(row=row_count, column=i+1).number_format = 'General'
                        except ValueError:
                            processed_row.append(value)
                    elif i == 1:  # 'Time' column (X-axis data)
                        try:
                            num_value = float(value)
                            processed_row.append(num_value)
                            # Explicitly set number format for Time column
                            ws.cell(row=row_count, column=i+1).number_format = '0.000' # Three decimal places
                        except ValueError:
                            processed_row.append(value)
                    else:  # Vrms channels (Y-axis data)
                        try:
                            num_value = float(value)
                            processed_row.append(num_value)
                            # Explicitly set number format for Vrms columns
                            ws.cell(row=row_count, column=i+1).number_format = '0.000' # Three decimal places
                        except ValueError:
                            processed_row.append(None) # Append None for invalid Vrms values

                ws.append(processed_row) # Append the processed row after converting values
        print("Data successfully written to Excel worksheet and converted to numbers.")


        # --- Charting Section ---
        chart = ScatterChart()
        chart.title = "Vrms Over Time"
        chart.style = 10
        chart.x_axis.title = "Time (seconds)"
        chart.y_axis.title = "Vrms"
        max_row = ws.max_row
        # Start from column 3 (index 2 in 0-based) for Vrms_CH1
        for i in range(num_channels):
            # X-axis: Time (column B, index 1)
            x_values = Reference(ws, min_col=2, min_row=2, max_row=max_row)
            # Y-axis: Vrms for the current channel (starting from column C, index 2)
            y_values = Reference(ws, min_col=3 + i, min_row=2, max_row=max_row)
            series = Series(y_values, x_values, title=f"Vrms_CH{i+1}")
            chart.series.append(series)
        # Add the chart to the worksheet
        ws.add_chart(chart, "E2") # Adjust cell to place the chart as needed
        # --- End Charting Section ---

        wb.save(full_excel_path)
        print(f"Excel data and chart saved successfully to: {full_excel_path}")

    except FileNotFoundError:
        print(f"Error: CSV data file not found at {full_csv_path}. Cannot create Excel file.")
    except Exception as e:
        print(f"An error occurred while creating the Excel file: {e}")

# --- MAIN ---
if __name__ == "__main__":
    # Initialize the Resource Manager
    rm = pyvisa.ResourceManager('@py')
    print("Resources found " , rm.list_resources())

    sample_time = sample_period(MIN_ACQUISITION_INTERVAL)

    # Create Event object for sampling time and start timer
    acquisition_allowed_event = threading.Event()
    timer_thread = threading.Thread(
        target=timer_thread_func,
        args=(acquisition_allowed_event, sample_time, stop_program_event),
        daemon=True # Daemon threads exit automatically when the main program exits
    )
    timer_thread.start()

    # Register the 'q' hotkey
    keyboard.add_hotkey('q', on_q_press)

    num_channels_to_monitor = 0
    connected_instrument = None
    datafile_name = None # Initialize datafile_name to None
    try:
        # Call the new function to connect to the instrument
        connected_instrument = connect_to_instrument(rm, DEFAULT_IP_ADDRESS)

        # Get the number of channels from the user
        num_channels_to_monitor = get_num_channels()

        # Set up channels based on the user input
        setup_needed = input("If answering no, I will attempt to setup contiguous channels (CH1 throug X) and measurements.  Leave scope alone (y/n)? :").strip()
        if setup_needed.lower() == 'y':
            print("Skipping scope setup. Ensure channels are configured correctly before starting data acquisition.")
        else:
            setup_scope(connected_instrument, num_channels_to_monitor)            

        # Create a data file for logging
        starting_date_and_time = datetime.datetime.now()
        paths = make_datafile(num_channels_to_monitor, starting_date_and_time, DESKTOP)
        user_path = paths[0]
        datafile_name = paths[1]
        full_data_path = os.path.join(user_path, datafile_name)
        print("Created file for data as ", datafile_name)
        
        count = 1
        print("Press 'q' to stop the program at any time.")
        # Main loop
        while not stop_program_event.is_set():
            # Wait for the minimum acquisition interval to pass before arming
            acquisition_allowed_event.wait(timeout=0.1) # This will block until the timer thread sets the event
            if stop_program_event.is_set(): # Check immediately after waiting
                break
            if not acquisition_allowed_event.is_set():
                # Continue waiting for the interval to pass or q to be pressed.
                continue
            acquisition_allowed_event.clear() # Reset the event for the next cycle

            # OK to sample
            current_date_and_time = datetime.datetime.now()
            dt = current_date_and_time - starting_date_and_time
            dt_in_seconds = dt.total_seconds()

            v_rms_readings = []
            for i in range(1, num_channels_to_monitor + 1):
                try:
                    v_rms = float(connected_instrument.query(f"MEASUrement:MEAS{i}:VALue?"))
                    # Apply bounds using the new routine call
                    v_rms = apply_vrms_bounds(v_rms)
                    v_rms_readings.append(v_rms)
                except pyvisa.errors.VisaIOError as e:
                    print(f"Error reading RMS for Channel {i}: {e}. Skipping this channel for this sample.")
                    v_rms_readings.append(float('NAN')) # Append NaN if reading fails
                except ValueError:
                    print(f"Could not convert RMS reading for Channel {i} to float. Skipping.")
                    v_rms_readings.append(float('NAN'))

            add_sample_to_file(user_path, datafile_name, count, dt_in_seconds, v_rms_readings)
            
            # Print the current sample data
            print_output = f"Sample {count:4d},   Time: {dt_in_seconds:9.3f} sec"
            for i, v_rms in enumerate(v_rms_readings):
                print_output += f", CH{i+1}: {v_rms:6.3f}"
            print(print_output)
            
            # Increment the counter
            count += 1

            # Check if 'q' was pressed taking samples 
            if stop_program_event.is_set():
                # If 'q' was pressed, stop the acquisition on the scope
                connected_instrument.write("CLEAR")
                break

    except Exception as e:
        print(f"An error occurred during program execution: {e}")
    finally:
        # Always close the instrument connection and resource manager
        if 'connected_instrument' in locals() and connected_instrument:
            print("Closing instrument connection.")
            connected_instrument.close()
        if rm:
            print("Closing Resource Manager.")
            rm.close()
        
        # After data acquisition stops, write to Excel if a datafile was created
        if datafile_name and num_channels_to_monitor > 0:
            write_to_excel_with_chart(datafile_name, user_path, num_channels_to_monitor)
