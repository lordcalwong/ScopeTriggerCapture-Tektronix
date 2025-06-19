# Power Monitoring
#
# Continuous monitor Amplifier Output channels and/or Line Inputs with
# asynchronous ON and OFF time logging. Configure all measurements on scope
# to be # RMS readings (~10V/div and ~trigger level of 10V).
# Trigger level is 18W/ch (9.8V) ON and 1V/ch OFF.
# autorun.
# Keeps track of number of valid cycles.
# Uses pyvisa for generic scope SCPI communications.
#
# Author: C. Wong 2025XXXX

import time
import datetime
import os
import keyboard
import pyvisa
import threading
import openpyxl
from openpyxl.chart import ScatterChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.chart.series import Series
from openpyxl.chart.shapes import GraphicalProperties
# from openpyxl.styles.fills import PatternFill

# Configure IP '192.168.1.53', '10.101.100.151', '10.101.100.236', '10.101.100.254', '10.101.100.176'
DEFAULT_IP_ADDRESS = '10.101.100.151'   # CHANGE FOR YOUR PARTICULAR SCOPE!
SAVE_PATH = r"C:\Users\Calvert.Wong\OneDrive - qsc.com\Desktop"
MIN_ACQUISITION_INTERVAL = 1   # sampling rate
MAX_VRMS = 999; MIN_VRMS = 0.001

# Global flag to signal the main loop and threads to stop
stop_program_event = threading.Event()

# --- FUNCTION DEFINITION ---
def timer_thread_func(event_to_set: threading.Event, interval: float, stop_event: threading.Event):
    """
    Separate thread that signals an event after the specified interval.
    """
    while not stop_event.is_set():
        # Wait for the interval, but allow to be interrupted by stop_event
        # If stop_event is set during this wait, it will return True immediately.
        if stop_event.wait(interval):
            break # Exit the loop if stop_event was set
        # If the wait completed (meaning interval passed and stop_event wasn't set),
        # then set the event for the main thread.
        if not stop_event.is_set():
            event_to_set.set()

def on_q_press():
    """
    Callback function when 'q' is pressed.
    """
    print("\n'q' pressed. Signaling program to stop.")
    stop_program_event.set()

def connect_to_instrument(resource_manager: pyvisa.ResourceManager, default_ip: str = DEFAULT_IP_ADDRESS):
    """
    Prompts the user for an IP address and attempts to establish a connection
    to a PyVISA instrument, retrying until successful.

    Args:
        resource_manager: The PyVISA ResourceManager instance.
        default_ip: The default IP address to suggest to the user.

    Returns:
        The connected PyVISA instrument object.
    """
    my_instrument = None
    while my_instrument is None:
        ip_address_input = input(
            f"Enter the instrument's IP address (e.g., {default_ip}) or 'd' for default: "
        ).strip()

        if ip_address_input.lower() == 'd':
            visa_address = default_ip
        else:
            visa_address = ip_address_input

        # Construct the full VISA resource string
        resource_string = f'TCPIP::{visa_address}::INSTR'

        print(f"Attempting to connect to: {resource_string}")

        try:
            # Attempt to open the resource
            my_instrument = resource_manager.open_resource(resource_string)

            # Try to query the instrument to verify connection
            print(f"Successfully connected! Instrument ID: {my_instrument.query('*IDN?').strip()}")

        except pyvisa.errors.VisaIOError as e:
            print(f"Connection failed: {e}")
            print("Please ensure the IP address is correct and the instrument is on and connected to the network.")
            print("Retrying in 2 seconds...")
            my_instrument = None   # Ensure my_instrument is None to continue the loop
            time.sleep(2)   # Wait before retrying

        except Exception as e:
            print(f"An unexpected error occurred: {e}")
            print("Retrying in 2 seconds...")
            my_instrument = None
            time.sleep(2)

    return my_instrument

def get_num_channels():
    """
    Asks the user to input the number of channels (1-8).
    """
    while True:
        try:
            num_channels = int(input("Enter the number of channels to monitor (1-8): "))
            if 1 <= num_channels <= 8:
                return num_channels
            else:
                print("Invalid input. Please enter a number between 1 and 8.")
        except ValueError:
            print("Invalid input. Please enter a number.")

def get_max_graph_voltage():
    """
    Asks the user to input maximum expected voltage for plots, e.g., 15, 50, 100, 150 volts
    """
    while True:
        try:
            max_voltage = int(input("Enter the maximum voltage for plots: "))
            if 10 <= max_voltage <= 400:
                return max_voltage
            else:
                print("Invalid input. Please enter a number between 1 and 400.")
        except ValueError:
            print("Invalid input. Please enter a number.")

def setup_scope(scope_device, num_channels):
    """
    Configures channels for the specified number of channels.
    Minimally tries to set scale and position.
    """
    print("Setting up oscilloscope...", end='')

    # # Clear existing measurements to avoid conflicts
    # scope_device.write("MEASUrement:CLEarALL")

    for i in range(1, num_channels + 1):
        scope_device.write(f"SELect:CH{i} ON")
        scope_device.write(f"CH{i}:SCALe 10")
        scope_device.write(f"CH{i}:POSition -4")
        # scope_device.write(f"MEASUrement:MEAS{i}:ENABle ON")
        # scope_device.write(f"MEASUrement:MEAS{i}:SOUrce1 CH{i}; STATE 1; TYPE RMS")  #works for DPO4K
        scope_device.write(f"MEASUrement:MEAS{i}:SOUrce CH{i}; TYPE RMS")  #works for MSO5

    # Wait for scope to finish setting up
    scope_device.query("*OPC?")
    print("Scope setup complete.")

def make_datafile(num_channels, timestamp):
    """Generates an Excel data file for the data based on start date and time.
    Includes headers for each monitored channel and sets column formatting.
    """
    excel_file_name = timestamp.strftime("%Y%m%d_%H%M%S.xlsx")
    full_data_path = os.path.join(SAVE_PATH, excel_file_name)

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Power Monitoring Data"

    header = ["Count", "Time"]
    for i in range(1, num_channels + 1):
        header.append(f"Vrms_CH{i}")
    sheet.append(header)

    # Set number format for Vrms columns
    # Vrms_CH1 is in column C (index 2 for 0-based, or 3 for 1-based)
    # So, columns from 3 up to 3 + num_channels - 1 will be Vrms
    for col_num in range(3, 3 + num_channels):
        sheet.column_dimensions[get_column_letter(col_num)].number_format = '0.000'

    print(f"Creating new Excel data file: {full_data_path}")
    return workbook, excel_file_name

def add_sample_to_file(workbook, data_file_name, counter, time_in_seconds, v_rms_values):
    """
    Appends sample data for all channels to the specified Excel file.
    """
    sheet = workbook.active
    row_data = [counter, time_in_seconds] + v_rms_values
    sheet.append(row_data)

def apply_vrms_bounds(v_rms):
    """
    Applies upper and lower bounds to the Vrms reading.
    """
    return max(min(v_rms, MAX_VRMS), MIN_VRMS)

def save_excel_with_chart(excel_workbook, datafile_name, num_channels_to_monitor, min_vrms_limit, max_vrms_limit, save_directory):
    """
    Saves the Excel workbook and adds a scatter chart to it.

    Args:
        excel_workbook: The openpyxl Workbook object.
        datafile_name: The name of the Excel file.
        num_channels_to_monitor: The number of channels being monitored.
        min_vrms_limit: The minimum limit for the Vrms y-axis.
        max_vrms_limit: The maximum limit for the Vrms y-axis.
        save_directory: The directory where the file should be saved.
    """
    if not excel_workbook or not datafile_name:
        print("No Excel workbook or data file name provided for saving.")
        return

    try:
        sheet = excel_workbook.active

        if sheet.max_row > 1: # Ensure there's data beyond the header for plotting
            chart = ScatterChart()
            chart.title = "Vrms vs. Time"
            chart.style = 13
            chart.x_axis.title = "Time (seconds)"
            chart.y_axis.title = "Vrms"
            # Set Y-axis limits
            chart.y_axis.scaling.min = min_vrms_limit
            chart.y_axis.scaling.max = max_vrms_limit

            # Set legend fill to white (using the updated method)
            # chart.legend.graphical_properties = GraphicalProperties(solidFill='FFFFFFFF')

            # Define the X-axis (Time) data
            times = Reference(sheet, min_col=2, min_row=2, max_row=sheet.max_row)

            for i in range(1, num_channels_to_monitor + 1):
                # Define the Y-axis (Vrms_CHx) data
                vrms_values = Reference(sheet, min_col=i + 2, min_row=2, max_row=sheet.max_row)
                series = Series(vrms_values, times, title=f"Vrms_CH{i}")
                # Set line to be connected point-to-point (no smoothing)
                # if series.graphicalProperties.line: # Ensure line object exists
                #     series.graphicalProperties.line.smooth = False
                # else: # Create line object if it doesn't exist to set properties
                #     from openpyxl.drawing.line import LineProperties
                #     series.graphicalProperties.line = LineProperties(smooth=False)
                chart.series.append(series)

            sheet.add_chart(chart, "B5")
            full_save_path = os.path.join(save_directory, datafile_name)
            print(f"Saving Excel file with chart: {full_save_path}")
            excel_workbook.save(full_save_path)
        else:
            full_save_path = os.path.join(save_directory, datafile_name)
            print("No data points collected to create a chart. Saving Excel file without chart.")
            excel_workbook.save(full_save_path)

    except Exception as e:
        print(f"Error saving Excel file or creating chart: {e}")


# --- MAIN ---
if __name__ == "__main__":
    # Initialize the Resource Manager
    rm = pyvisa.ResourceManager('@py')
    print(rm.list_resources())

    # Create Event object for sampling time and start timer
    acquisition_allowed_event = threading.Event()
    timer_thread = threading.Thread(
        target=timer_thread_func,
        args=(acquisition_allowed_event, MIN_ACQUISITION_INTERVAL, stop_program_event),
        daemon=True # Daemon threads exit automatically when the main program exits
    )
    timer_thread.start()

    # Register the 'q' hotkey
    keyboard.add_hotkey('q', on_q_press)

    num_channels_to_monitor = 0
    connected_instrument = None
    excel_workbook = None
    datafile_name = None
    try:
        # Call the new function to connect to the instrument
        connected_instrument = connect_to_instrument(rm, DEFAULT_IP_ADDRESS)

        # Get the number of channels from the user
        num_channels_to_monitor = get_num_channels()
        # Get maximum voltage for plots
        max_graph_voltage= get_max_graph_voltage()
        # Set up channels based on the user input
        setup_scope(connected_instrument, num_channels_to_monitor)

        # Create an Excel data file for logging
        starting_date_and_time = datetime.datetime.now()
        excel_workbook, datafile_name = make_datafile(num_channels_to_monitor, starting_date_and_time)
        print("Created file for data as ", datafile_name)

        count = 1
        # Main loop
        while not stop_program_event.is_set():
            # Wait for the minimum acquisition interval to pass before arming
            acquisition_allowed_event.wait(timeout=0.05)
            if stop_program_event.is_set():
                break
            if not acquisition_allowed_event.is_set():
                # Continue waiting for the interval to pass or q to be pressed.
                continue
            acquisition_allowed_event.clear() # Reset the event for the next cycle

            # OK to sample
            current_date_and_time = datetime.datetime.now()
            dt = current_date_and_time - starting_date_and_time
            dt_in_seconds = dt.total_seconds()

            v_rms_readings = []
            for i in range(1, num_channels_to_monitor + 1):
                try:
                    v_rms = float(connected_instrument.query(f"MEASUrement:MEAS{i}:VALue?"))
                    # Apply bounds using the new routine call
                    v_rms = apply_vrms_bounds(v_rms)
                    v_rms_readings.append(v_rms)
                except pyvisa.errors.VisaIOError as e:
                    print(f"Error reading RMS for Channel {i}: {e}. Skipping this channel for this sample.")
                    v_rms_readings.append(float('NAN')) # Append NaN if reading fails
                except ValueError:
                    print(f"Could not convert RMS reading for Channel {i} to float. Skipping.")
                    v_rms_readings.append(float('NAN'))

            add_sample_to_file(excel_workbook, datafile_name, count, dt_in_seconds, v_rms_readings)

            # Print the current sample data  (make comment out if fast sampling)
            print_output = f"Sample {count:4d}: Time: {dt_in_seconds:9.3f} sec"
            for i, v_rms in enumerate(v_rms_readings):
                print_output += f", CH{i+1}: {v_rms:3.3f}"
            print(print_output)

            # Increment the counter
            count += 1

            # Check if 'q' was pressed taking samples
            if stop_program_event.is_set():
                # If 'q' was pressed, stop the acquisition on the scope
                connected_instrument.write("CLEAR")
                break

    except Exception as e:
        print(f"An error occurred during program execution: {e}")
    finally:
        # Always close the instrument connection and resource manager
        if 'connected_instrument' in locals() and connected_instrument:
            print("Closing instrument connection.")
            connected_instrument.close()
        if rm:
            print("Closing Resource Manager.")
            rm.close()

        # Save the Excel workbook and add chart
        save_excel_with_chart(excel_workbook, datafile_name, num_channels_to_monitor, MIN_VRMS, max_graph_voltage, SAVE_PATH)
