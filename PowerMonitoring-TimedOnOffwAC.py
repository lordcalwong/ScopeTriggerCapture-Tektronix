# Power Monitoring- Triggered
#
# Continuously monitor amplifier output channels (CH2+) and an AC Line
# monitor (CH1) to log ON/OFF times (with better than 1 second resolution).
#
# Scope trigger must be autorun. Code assumes system will  starts # in an
# "OFF" state.
#
# User inputs IP address and number of channels to monitor (1-8), and
# the option to set up scope or allow contiguous channels to be configured
# for RMS measurements.  User sets desired ON or OFF threshold.
#
# Saves data to csv file to user path or defaults to the desktop.
#
# Author: C. Wong 20250728

import time
import datetime
import os
import pyvisa
import threading
import csv
import keyboard
from collections import deque  #only needed for running average on AC Line

from openpyxl import Workbook

DEFAULT_IP_ADDRESS = '192.168.1.53'  #default IP, 192.168.1.53, 10.101.100.151
MAX_LINE_VOLTAGE_VRMS = 350.0  # AC Line RMS limit, arbitrary limit for CH1
MAX_VRMS = 50  # ~312 W arbitrary limit for other audio CHs  (CH2+)
ON_THRESHOLD = 9.0  #default trigger levels for 'ON' per audio CH
OFF_THRESHOLD = 2.0 #default trigger levels for 'OFF' per audio CH
LINE_VOLTAGE_WINDOW_SIZE = 4 # Define the window size for the running average

# Find user desktop one level down from home [~/* /Desktop] as optional path
from glob import glob
DESKTOP = glob(os.path.expanduser("~\\*\\Desktop"))
# If not found, revert to the standard location (~/Desktop)
if len(DESKTOP) == 0:
    DESKTOP = os.path.expanduser("~\\Desktop")
else:
    DESKTOP = DESKTOP[0] # glob returns a list, take the first result

# Global flag to signal the main loop and threads to stop
stop_program_event = threading.Event()

def on_q_press():
    """
    Callback function when 'q' is pressed.
    """
    print("\n'q' pressed. Signaling program to stop.")
    stop_program_event.set()

def on_esc_press():
    """
    Callback function when 'Esc' is pressed.
    """
    print("You pressed Esc!")
    stop_program_event.set()

def connect_to_instrument(resource_manager: pyvisa.ResourceManager, default_ip: str = DEFAULT_IP_ADDRESS):
    """
    Prompts the user for an IP address and attempts to establish a connection
    to a PyVISA instrument, retrying until successful.

    Args:
        resource_manager: The PyVISA ResourceManager instance.
        default_ip: The default IP address to suggest to the user.

    Returns:
        The connected PyVISA instrument object.
    """
    my_instrument = None
    while my_instrument is None:
        ip_address_input = input(
            f"Enter the instrument's IP address or 'd' for default ({default_ip}): "
        ).strip()
        if ip_address_input.lower() == 'd':
            visa_address = default_ip
        else:
            visa_address = ip_address_input

        # Construct the full VISA resource string
        resource_string = f'TCPIP::{visa_address}::INSTR'

        print(f"Attempting to connect to: {resource_string}")

        try:
            # Attempt to open the resource
            my_instrument = resource_manager.open_resource(resource_string)

            # Try to query the instrument to verify connection
            print(f"Successfully connected! Instrument ID: {my_instrument.query('*IDN?').strip()}")

        except pyvisa.errors.VisaIOError as e:
            print(f"Connection failed: {e}")
            print("Please ensure the IP address is correct and the instrument is on and connected to the network.")
            print("Retrying in 2 seconds...")
            my_instrument = None   # Ensure my_instrument is None to continue the loop
            time.sleep(2)   # Wait before retrying

        except Exception as e:
            print(f"An unexpected error occurred: {e}")
            print("Retrying in 2 seconds...")
            my_instrument = None
            time.sleep(2)

    return my_instrument

def get_num_channels():
    """
    Asks the user to input the number of channels (2-8).
    First channel monitors AC Line. Additional channels monitor
    amplifier outputs (CH2+).
    """
    print("First channel monitors AC Line. Additional channels (CH2+) monitor amplifier outputs.")
    while True:
        try:
            num_channels = int(input("Including AC Line (CH1), enter no. of CHs (2-8): "))
            if 2 <= num_channels <= 8:
                return num_channels
            else:
                print("Invalid input. Please enter a number between 2 and 8.")
        except ValueError:
            print("Invalid input. Please enter a number.")

def get_thresholds(default_on_trig: float = ON_THRESHOLD, default_off_trig: float = OFF_THRESHOLD):
    """
    Prompts the user for ON and OFF threshold levels for Vrms,
    validating the inputs.
    """
    on_trig_level = default_on_trig
    off_trig_level = default_off_trig

    while True:
        try:
            on_trig_input = input(f"Enter trigger level for ON cycle (default: {on_trig_level:.2f}V, 'd' for default): ").strip()
            if on_trig_input.lower() == 'd':
                on_trig_level = default_on_trig
            else:
                on_trig_level = float(on_trig_input)

            off_trig_input = input(f"Enter trigger level for OFF cycle (default: {off_trig_level:.2f}V, 'd' for default): ").strip()
            if off_trig_input.lower() == 'd':
                off_trig_level = default_off_trig
            else:
                off_trig_level = float(off_trig_input)

            # Validate the thresholds
            if on_trig_level <= 0:
                print("Error: ON threshold must be greater than 0.")
            elif off_trig_level <= 0:
                print("Error: OFF threshold must be greater than 0.")
            elif on_trig_level >= MAX_VRMS:
                print(f"Error: ON threshold must be less than {MAX_VRMS}V.")
            elif off_trig_level >= on_trig_level:
                print("Error: OFF threshold must be less than the ON threshold.")
            else:
                return on_trig_level, off_trig_level # Valid inputs, exit loop

        except ValueError:
            print("Invalid input. Please enter a numerical value or 'd' for default.")
        except Exception as e:
            print(f"An unexpected error occurred: {e}")

def setup_scope(scope_device, num_channels):
    """
    Configures channels for the specified number of channels.
    Minimally tries to set scale and position.
    """
    print("Ok. Setting up oscilloscope...  ")

    scope_device.write("*RST")  # Only needed for stubborn scopes

   # Set up vertical
    for i in range(1, num_channels + 1):
        scope_device.write(f"SELect:CH{i} ON")
        scope_device.write(f"CH{i}:POSition 0")
        if i == 1:
            scope_device.write(f"CH{1}:SCALe 100")
        else:
            scope_device.write(f"CH{i}:SCALe 10")
        scope_device.write(f"MEASUrement:MEAS{i}:SOUrce CH{i}; STATE 1")   # Need separate STATE 1 command for DPO
        scope_device.write(f"MEASUrement:MEAS{i}:SOUrce CH{i}; TYPE RMS")  # Need separate TYPE command for MSO
        # scope_device.write(f"DISplay:SPECView1:VIEWStyle OVERly")   # Not working on MSO

   # Set up timebase
    scope_device.write("HORizontal:SCAle 200E-6")
    scope_device.write("HORizontal:POSition 50")

    # Set up trigger to line
    scope_device.write("TRIGger:A:EDGE:SOUrce CH1")
    scope_device.write("TRIGger:A:EDGE:COUPling DC")
    scope_device.write("TRIGger:A:EDGE:SLOpe RISE")
    scope_device.write("TRIGger:A:LEVel:CH1 50")

    print("Scope setup complete.")
    print("Check and adjust scale, timing, and trigger is as needed. ")

    # # Wait for scope to finish setting up
    # scope_device.query("*OPC?")  # Issue with MSO
    time.sleep(0.2)

def make_datafile(timestamp, desktoppath: str = DESKTOP): # num_channels removed as it's not needed for header anymore
    """Generates a data file for the data based on start date and time.
    Includes headers for each monitored channel. Asks user for directory
    or defaults to desktop.
    Returns tuple for user_path and data_log_file_name
    """
    while True:
        try:
            user_path_input = input(f"Enter path for data or 'd' for default ({desktoppath}): ").strip()
            if user_path_input.lower() == 'd':
                user_path = desktoppath
            else:
                user_path = user_path_input

            data_log_file_name = timestamp.strftime("%Y%m%d_%H%M%S.txt")
            full_data_path = os.path.join(user_path, data_log_file_name)

            # Check if data file already exists
            if not os.path.exists(full_data_path):
                print(f"Creating new data file at {user_path}.. .")
                with open(full_data_path, "w") as datafile:
                    # UPDATED: Header for duration logging
                    header = "Event_Count, Start_Time_Absolute, End_Time_Absolute, Line Voltage, State, Duration_Seconds"
                    datafile.write(header + "\n")
            else:
                print(f"File exist. We will be appending to existing file: {full_data_path}")
            return user_path, data_log_file_name

        except ValueError:
            print("Invalid path. Please enter a valid path.")

def apply_vrms_bounds(number: float) -> float:
    """
    Applies upper and lower bounds to the Vrms readings 2nd CH+, not the AC line (CH 1).
    """
    return max(min(number, MAX_VRMS), 0)

def apply_line_voltage_bounds(number: float) -> float:
    """
    Applies upper and lower bounds to the Vrms readings for the AC Line (CH1).
    """
    return max(min(number, MAX_LINE_VOLTAGE_VRMS), 0)

def log_duration_to_file(save_directory, data_file_name, event_count, start_time, end_time, line_voltage_avg, state, duration_seconds):
    """
    Appends duration data to the specified file.
    """
    try:
        datafile_and_path = os.path.join(save_directory, data_file_name)
        with open(datafile_and_path, "a") as f:
            line = f"{event_count:4d}, {start_time.strftime('%Y-%m-%d %H:%M:%S.%f')}, {end_time.strftime('%Y-%m-%d %H:%M:%S.%f')}, {line_voltage_avg:7.3f}, {state}, {duration_seconds:9.3f}"
            f.write(line + "\n")
    except IOError as e:
        print(f"Error appending data to file '{datafile_and_path}': {e}")

def write_to_excel(datafile_name: str, save_directory: str): # num_channels removed
    """
    Reads data from the specified CSV file, writes it to an Excel worksheet

    Args:
        datafile_name: The name of the CSV data file.
        save_directory: The directory where the CSV and Excel files are saved.
    """
    full_csv_path = os.path.join(save_directory, datafile_name)
    excel_file_name = os.path.splitext(datafile_name)[0] + ".xlsx"
    full_excel_path = os.path.join(save_directory, excel_file_name)

    print(f"\nAttempting to create Excel file: {full_excel_path}")

    wb = Workbook()
    ws = wb.active
    ws.title = "Power Monitoring Durations"

    try:
        with open(full_csv_path, 'r') as f:
            reader = csv.reader(f)
            # Write header row
            header = next(reader)
            ws.append(header)

            # Write data rows and convert to numbers, then format cells
            row_count = 1 # To track actual row in Excel, starting from 1 for headers
            for row in reader:
                row_count += 1 # Increment for data rows
                processed_row = []
                # Event_Count, Start_Time_Absolute, End_Time_Absolute, State, Duration_Seconds
                for i, value in enumerate(row):
                    if i == 0:  # Event_Count
                        try:
                            num_value = int(value)
                            processed_row.append(num_value)
                        except ValueError:
                            processed_row.append(value)
                    elif i in [1, 2]: # Start_Time_Absolute, End_Time_Absolute - keep as strings for now or convert to datetime objects if needed for excel, but direct plotting might be tricky
                        processed_row.append(value)
                    elif i == 3: # Line_Voltage
                        try:
                            num_value = float(value)
                            processed_row.append(num_value)
                            ws.cell(row=row_count, column=i+1).number_format = '0.000'
                        except ValueError:
                            processed_row.append(None)
                    elif i == 4: # State
                        processed_row.append(value)
                    elif i == 5: # Duration_Seconds
                        try:
                            num_value = float(value)
                            processed_row.append(num_value)
                            ws.cell(row=row_count, column=i+1).number_format = '0.000'
                        except ValueError:
                            processed_row.append(None) # Append None for invalid duration values
                ws.append(processed_row)
        wb.save(full_excel_path)
        print(f"Excel data saved successfully to: {full_excel_path}")

    except FileNotFoundError:
        print(f"Error: CSV data file not found at {full_csv_path}. Cannot create Excel file.")
    except Exception as e:
        print(f"An error occurred while creating the Excel file: {e}")

# ************** MAIN
rm = None
connected_instrument = None
last_state_change_time = datetime.datetime.now()
event_counter = 0
first_transition_logged = False

# Initialize deque for line voltage readings
line_voltage_readings_queue = deque(maxlen=LINE_VOLTAGE_WINDOW_SIZE)
# current_line_voltage_avg = 0.0 # This variable is no longer strictly needed as a global if using local snapshot
# It's better to calculate and use a local snapshot for current readings.

try:
    num_channels_to_monitor = 0
    connected_instrument = None
    datafile_name = None  # Initialize datafile_name to None
    current_state = "UNKNOWN"  # States can be "ON" or "OFF"
    # previous_state = current_state  # This variable is not used and can be removed for clarity

    # Register the 'q' hotkey
    keyboard.add_hotkey('q', on_q_press)
    keyboard.add_hotkey('esc', on_esc_press)

    # Initialize Resource Manager
    rm = pyvisa.ResourceManager()
    connected_instrument = connect_to_instrument(rm, DEFAULT_IP_ADDRESS)
    if connected_instrument is None:
        print("Failed to connect to the instrument. Exiting.")
        exit() # Exit if connection failed

    # Get the number of channels from user
    num_channels_to_monitor = get_num_channels()

    # Get desired ON/OFF thresholds from user
    limits = get_thresholds()
    high_limit = limits[0]
    low_limit = limits[1]
    print("high_limit = ", high_limit, ", low_limit = ", low_limit)

    # Set up channels if needed based on the user input
    setup_needed = input("(L)eave scope alone or (S)etup contiguous channels?: ").strip()
    if setup_needed.lower() == 'l':
        print("Skipping scope setup. Ensure channels are configured correctly before starting data acquisition.")
    else:
        setup_scope(connected_instrument, num_channels_to_monitor)

    # Create a data file for logging
    last_state_change_time = datetime.datetime.now()
    paths = make_datafile(last_state_change_time, DESKTOP)
    user_path = paths[0]
    datafile_name = paths[1]
    full_data_path = os.path.join(user_path, datafile_name)
    print("Created file for data as", datafile_name)

    # Notify user off and running
    print(f"Monitoring for ON (all channels > {high_limit:.2f} Vrms) and OFF (all channels < {low_limit:.2f} Vrms) states.")
    no_response = input("Check Line voltage and amp out are OFF (0V) before begining. Hit Enter to continue...")
    print("Press 'q','esc', or 'Crtl-C' to stop the program at any time.")
    print("Start monitoring...")

    # Main loop
    while not stop_program_event.is_set():
        time.sleep(0.05) # Small delay for keyboard input before checking if scope triggered

        # Check scope in run mode
        connected_instrument.write("ACQuire:STATE ON")

        # Read measurements
        current_time = datetime.datetime.now()
        v_rms_readings = []

        # Initialize current_line_voltage_snapshot for this loop iteration
        # This will be updated if line_voltage_readings_queue is populated.
        current_line_voltage_snapshot = 0.0 # Initialize to 0.0 before reading CH1

        for i in range(1, num_channels_to_monitor + 1):
            try:
                v_rms = float(connected_instrument.query(f"MEASUrement:MEAS{i}:VALue?"))
                if i == 1: # This is the AC Line (CH1)
                    v_rms = apply_line_voltage_bounds(v_rms)
                    line_voltage_readings_queue.append(v_rms)
                    if len(line_voltage_readings_queue) > 0:
                        current_line_voltage_snapshot = sum(line_voltage_readings_queue) / len(line_voltage_readings_queue)
                    # Else, current_line_voltage_snapshot remains 0.0 if queue is empty (shouldn't happen here)
                else: # Apply bounds only if it's NOT channel 1
                    v_rms = apply_vrms_bounds(v_rms)
                v_rms_readings.append(v_rms)

            except pyvisa.errors.VisaIOError as e:
                print(f"Error reading RMS for Channel {i}: {e}. Skipping this channel for this sample.")
                v_rms_readings.append(float('NAN')) # Append NaN if reading fails

            except ValueError:
                print(f"Could not convert RMS reading for Channel {i} to float. Skipping.")
                v_rms_readings.append(float('NAN'))

        # Create a new list for checks, excluding channel 1
        v_rms_readings_for_state_check = v_rms_readings[1:] # Slice from the second element to the end

        # Check if any readings are NaN, if so, we can't determine state reliably
        if any(v == float('NAN') for v in v_rms_readings):
            print("Warning: Skipping state evaluation due to invalid Vrms readings.")
            continue

        # Determine the potential new state based on current readings for CH2+
        all_channels_on = all(v >= high_limit for v in v_rms_readings_for_state_check)
        all_channels_off = all(v <= low_limit for v in v_rms_readings_for_state_check)

        # State Establishment/Transition Logic
        if current_state == "UNKNOWN":
            if all_channels_on:
                current_state = "ON"
                last_state_change_time = current_time
                connected_instrument.write(f"TRIGger:A:LEVel:CH1 {low_limit}")
                print(f"Initial state detected as ON. Will start logging on next transition to OFF.")
            elif all_channels_off:
                current_state = "OFF"
                last_state_change_time = current_time
                connected_instrument.write(f"TRIGger:A:LEVel:CH1 {high_limit}")
                print(f"Initial state detected as OFF. Will start logging on next transition to ON.")
            else:
                # Still in an indeterminate state or no clear ON/OFF. Keep current_state as UNKNOWN.
                pass # No change, continue will be called below
            continue # Always continue if still in UNKNOWN or just established initial state

        # Current_state is either 'ON' or 'OFF'. Proceed to check for transitions.
        new_actual_state = current_state # Assume no change unless clear transition

        # Determine if there's a *real* transition from the established state
        if current_state == "OFF" and all_channels_on:
            new_actual_state = "ON"
        elif current_state == "ON" and all_channels_off:
            new_actual_state = "OFF"
        # If both all_channels_on and all_channels_off are false, new_actual_state remains current_state
        # This handles the intermediate or stable state where no clear transition occurs.

        if new_actual_state != current_state:
            if not first_transition_logged:
                # This is the very first transition detected. Don't log the initial state.
                print(f"State captured. Waiting for first transition to ON.")
                current_state = new_actual_state
                last_state_change_time = current_time
                first_transition_logged = True
                # Set the trigger level for the next transition
                if current_state == "ON":
                    connected_instrument.write(f"TRIGger:A:LEVel:CH1 {low_limit}")
                else:
                    connected_instrument.write(f"TRIGger:A:LEVel:CH1 {high_limit}")
            else: 
                # This is a subsequent transition; start logging from this point
                duration = (current_time - last_state_change_time).total_seconds()
                event_counter += 1
                if current_state == "OFF" and new_actual_state == "ON":
                    # Transition from OFF to ON
                    # Log previous OFF state with 0.0 line voltage
                    log_duration_to_file(user_path, datafile_name, event_counter, last_state_change_time, current_time, 0.0, "OFF", duration)
                    print(f"State Change: OFF to ON. Previous OFF duration: {duration:.3f} seconds.")
                    current_state = "ON"
                    last_state_change_time = current_time
                    connected_instrument.write(f"TRIGger:A:LEVel:CH1 {low_limit}") # Set trigger for next OFF detection

                elif current_state == "ON" and new_actual_state == "OFF":
                    # Transition from ON to OFF
                    # Determine the line voltage to log
                    voltage_to_log = current_line_voltage_snapshot

                    # HARD-CODE: Set line voltage to 0.0 if this is the very first ON-to-OFF transition (event_counter == 1)
                    # This explicitly handles the initial corrupted reading you observed.
                    if event_counter == 1:
                        voltage_to_log = 0.0

                    log_duration_to_file(user_path, datafile_name, event_counter, last_state_change_time, current_time, voltage_to_log, "ON", duration)
                    print(f"State Change: ON to OFF. Previous ON duration: {duration:.3f} seconds. Line Voltage: {voltage_to_log:.3f}Vrms")
                    current_state = "OFF"
                    last_state_change_time = current_time
                    connected_instrument.write(f"TRIGger:A:LEVel:CH1 {high_limit}") # Set trigger for next ON detection

            # # debug
            # print(f"CH2+ Readings for State Check: {[f'{v:.3f}' for v in v_rms_readings_for_state_check]}")
            # print(f"All channels ON condition: {all_channels_on}")
            # print(f"All channels OFF condition: {all_channels_off}")
            # print(f"Current State: {current_state}")

except KeyboardInterrupt:
    print("\nProgram terminated by user (Ctrl+C).")
except Exception as e:
    print(f"An error occurred during program execution: {e}")
finally:
    # Always close the instrument connection and resource manager
    if 'connected_instrument' in locals() and connected_instrument:
        try:
            connected_instrument.write("ACQuire:STATE OFF") # Stop acquisition before closing
            connected_instrument.write("CLEAR") # Ensure scope acquisition is stopped
            connected_instrument.close()
            print("Instrument connection closed.")
        except pyvisa.errors.VisaIOError as e:
            print(f"Error closing instrument connection: {e}")
    if rm:
        try:
            rm.close()
            print("Resource Manager closed.")
        except Exception as e:
            print(f"Error closing Resource Manager: {e}")

    # Before exiting, log the duration of the final state if it was not already logged
    final_time = datetime.datetime.now()
    duration = (final_time - last_state_change_time).total_seconds()

    # Determine the line voltage to log for the final state
    final_voltage_to_log = 0.0 # Default to 0.0

    # Only attempt to use current_line_voltage_snapshot if the queue has data
    if len(line_voltage_readings_queue) > 0:
        final_voltage_to_log = sum(line_voltage_readings_queue) / len(line_voltage_readings_queue)

    if current_state == "OFF":
        # Always log 0.0 for line voltage if the final state is OFF
        log_duration_to_file(user_path, datafile_name, event_counter, last_state_change_time, final_time, 0.0, current_state, duration)
        print(f"Program stopped. Final {current_state} duration: {duration:.3f} seconds. Line Voltage: 0.000Vrms (Hardcoded for OFF state)")
    else: # If the final state was ON
        # Apply the same hard-code logic for the final ON state if event_counter is 0 or 1
        # This handles cases where the program exits very quickly after starting
        if event_counter == 0 or (event_counter == 1 and duration < 1.0): # event_counter 0 means no transitions logged. 1 means the initial bad one was.
             final_voltage_to_log = 0.0 # Hardcode if it's very early in program execution or only initial ON
             print("Note: Line voltage for final ON state hard-coded to 0.0V due to early program termination or initial state.")

        log_duration_to_file(user_path, datafile_name, event_counter, last_state_change_time, final_time, final_voltage_to_log, current_state, duration)
        print(f"Program stopped. Final {current_state} duration: {duration:.3f} seconds. Line Voltage: {final_voltage_to_log:.3f}Vrms")
        